# Exporting data from existing initial database files
# into exel files matchin the to[_template

import copy
from openpyxl import load_workbook
import openpyxl.utils.cell
import shutil
from utils import *

import sys



TIPS_DIR = BASE_DIR+'/tips'


tips = []
for LANG in LANGS:
    print("### Extrating tips for", LANG)
    filepath = TIPS_DIR+'/'+LANG+'/tips.xlsx'

    workbook = load_workbook(filepath)
    trait_name = None
    for READER_TYPE in READER_TYPES:
        # Load proper sheet
        sheet = workbook[READER_TYPE]
        
        for i in range(0,10000):
            row = 5+(i*4)
            key  = sheet.cell(row=row, column=2).value
            if not key: 
                print('{0: <5}'.format(i), "tips for", LANG, "->", READER_TYPE)
                break
            
            ageGroup = sheet.cell(row=row, column=3).value
            # Check if tip was already created for other language
            isNew = True
            try: # Find existing object
                tip = next(t for t in tips if t['key']==key and t['readerType']==READER_TYPE and t['ageGroup'] == ageGroup)
                isNew = False
            except:# Create object
                tip = copy.deepcopy(TIP_SCHEMA)
                tip['key'] = key
                tip['readerType'] = READER_TYPE
                tip['ageGroup'] = ageGroup
            tip['introduction'][LANG] = sheet.cell(row=row, column=4).value
            tip['text'][LANG] = sheet.cell(row=row, column=7).value
            tip['reasoning'][LANG] = sheet.cell(row=row, column=10).value
            
            # If not on the list
            if (isNew): tips.append(tip)

# Update initial database files
with open(INIT_DATABASE_PATH+"/cognitiveTips.js", "w") as output:
    # Sort
    tips.sort(key = lambda x: int(x['key'].split('_')[0]))

    
    tips.sort(key = lambda x: int(x['ageGroup']))
    tips.sort(key = lambda x: int(x['key'].split('_')[1]))
    order = ['student', 'teacher', 'parent', 'employee', 'leader', 'team-leader', 'class-teacher']
    tips.sort(key = lambda x: order.index(x['readerType']))
    
    data = json.dumps(tips, indent=4)
    output.write("exports.cognitiveTips = "+data)
            