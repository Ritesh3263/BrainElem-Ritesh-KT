# Importing data from exel files into initial database files
# It will also update translations files on frontent containing names of the tratis



#pip install styleframe
# pip install pandas

import os, copy
from openpyxl import load_workbook
import openpyxl.utils.cell
from utils import *



TRAITS_DIR = BASE_DIR+'/traits'

traits = {}

# This object will be used to update translations on the frontent
frontend_translations = {}
for LANG in LANGS: frontend_translations[LANG]={}

# transform value to empty string if it's a null
def not_null(value):
    return value.strip() if value else ""

for LANG in LANGS:
    print("### Processing files for", LANG)
    dirs = os.listdir(TRAITS_DIR+'/'+LANG)
    files = []
    for dir_name in dirs:
        dir_path = TRAITS_DIR+'/'+LANG+"/"+dir_name
        files_in_dir = [dir_path+'/'+f for f in os.listdir(dir_path)]
        files+=files_in_dir

    # Sort
    files.sort(key = lambda x: x.split('[id')[1].replace(':', '').replace('_', ''))

    # For testing to switch order
    #myorder = [0, 3, 4, 1, 2, 5]
    #files = [files[i] for i in myorder]
    for filepath in files:
        workbook = load_workbook(filepath)
        trait_name = None
        for READER_TYPE in READER_TYPES:
            # Load proper sheet
            sheet = workbook[READER_TYPE]
            if (not trait_name): 
                trait_name = sheet.cell(row=2, column=2).value
                print('-->',trait_name)
            # Create initial schema
            if (READER_TYPE+"->"+trait_name not in traits):
                    traits[READER_TYPE+"->"+trait_name] = copy.deepcopy(TRAIT_SCHEMA)
                    traits[READER_TYPE+"->"+trait_name]['key'] = trait_name
                    traits[READER_TYPE+"->"+trait_name]['readerType'] = READER_TYPE
            ##############################
            ####  BASIC DATA #############
            ##############################
            traits[READER_TYPE+"->"+trait_name]['key'] = trait_name
            traits[READER_TYPE+"->"+trait_name]['readerType'] = READER_TYPE
            traits[READER_TYPE+"->"+trait_name]['abbreviation'][LANG] = not_null(sheet.cell(row=3, column=2).value)
            short_name = not_null(sheet.cell(row=3, column=3).value)
            traits[READER_TYPE+"->"+trait_name]['shortName'][LANG] = short_name
            traits[READER_TYPE+"->"+trait_name]['shortDescription'][LANG] = not_null(sheet.cell(row=6, column=2).value) 
            traits[READER_TYPE+"->"+trait_name]['mainDefinition'][LANG] = not_null(sheet.cell(row=7, column=2).value)
            
            # This object will be used to update translations on the frontent
            key = trait_name+'-short-name'
            if (short_name and key not in frontend_translations[LANG]):
                if 'strong-and-weak-points-for' not in key:
                    frontend_translations[LANG][key] = short_name
            #############################
            #### PERSONALIZED TEXTS #####
            #############################
            for level in [1,2,3,4,5]:
                level_label = 'level_'+str(level)
                for index in [0,1,2,3,4]:
                    #####################
                    #### DESCRIPTIONS ###
                    #####################
                    row = 9+level
                    column = 2+3*index
                    text = sheet.cell(row=row, column=column).value
                    if (text and text.strip()): 
                        traits[READER_TYPE+"->"+trait_name]['descriptions'][level_label][LANG].append(text)
                    #####################
                    #### ACTIONS ########
                    #####################
                    row = 16+level
                    column = 2+3*index
                    text = sheet.cell(row=row, column=column).value
                    if (text and text.strip()):
                        traits[READER_TYPE+"->"+trait_name]['actions'][level_label][LANG].append(text)
            #############################
            #### LOWEST/HIGHEST #########
            #############################
            highest_def = sheet.cell(row=24, column=2).value
            traits[READER_TYPE+"->"+trait_name]['highestDefinition'][LANG] = highest_def if highest_def else ""
            lowest_def = sheet.cell(row=28, column=2).value
            traits[READER_TYPE+"->"+trait_name]['lowestDefinition'][LANG] = lowest_def if lowest_def else ""
            for index in [0,1,2,3,4]:
                column = 2+3*index
                # HIGHEST
                row = 25
                text = sheet.cell(row=row, column=column).value
                if (text and text.strip()):
                    traits[READER_TYPE+"->"+trait_name]['actions']['highest'][LANG].append(text)
                # LOWEST
                row = 29
                text = sheet.cell(row=row, column=column).value
                if (text and text.strip()):
                    traits[READER_TYPE+"->"+trait_name]['actions']['lowest'][LANG].append(text)
                # NEUROBIOLOGICAL EFFECTS
                row = 32
                text = sheet.cell(row=row, column=column).value
                if (text and text.strip()):
                    traits[READER_TYPE+"->"+trait_name]['neurobiologicalEffects'][LANG].append(text)


# Update frontent translations for traits
for LANG in LANGS:
    path = "./services/frontend/public/locales/"+LANG.lower()+"/traits.json"
    with open(path, "w") as output:
        data = json.dumps(frontend_translations[LANG], indent=2)
        output.write(data)
# Update initial database files
with open(INIT_DATABASE_PATH+"/cognitiveTraits.js", "w") as output:
    data = json.dumps([trait for trait in traits.values()], indent=2)
    output.write("exports.cognitiveTraits = "+data)
            


