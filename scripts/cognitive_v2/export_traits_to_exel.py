# Exporting data from existing initial database files
# into exel files matchin the trait_template

import copy
from openpyxl import load_workbook
import openpyxl.utils.cell
import shutil
from utils import *

TRAITS_DIR = BASE_DIR+'/traits'
recreate_dir(TRAITS_DIR)

# OLD # Traits in initial database files are divided into 3 main files
#grouped_traits = {} # Group traits from all those files into single object
# cognitiveTraitsForMySpace = load_json_from_initial_database_file('cognitiveTraitsForMySpace.js')
# cognitiveTraitsForReport = load_json_from_initial_database_file('cognitiveTraitsForReport.js')
# cognitiveTraitsForTeamModule = load_json_from_initial_database_file('cognitiveTraitsForTeamModule.js')
# for json_data in [cognitiveTraitsForMySpace, cognitiveTraitsForReport, cognitiveTraitsForTeamModule]:
#     for trait in json_data:
#         try: grouped_traits[trait['key']].append(trait)
#         except: grouped_traits[trait['key']] = [trait]

# NEW
grouped_traits = {}
cognitiveTraits = load_json_from_initial_database_file('cognitiveTraits.js')
for trait in cognitiveTraits:
    try: grouped_traits[trait['key']].append(trait)
    except: grouped_traits[trait['key']] = [trait]
  

# # Add missing traits
# names = {
# }
# for trait_key in [n.replace('-short-name', '') for n in names['EN'].keys()]:
#     if trait_key not in grouped_traits:
#         print("Missing", trait_key)
#         for READER_TYPE in READER_TYPES:
#             missing = copy.deepcopy(TRAIT_SCHEMA)
#             missing['readerType'] = READER_TYPE
#             missing['shortName'] = {
#                 "PL": names['PL'][trait_key+'-short-name'],
#                 "EN": names['EN'][trait_key+'-short-name'],
#                 "FR": names['FR'][trait_key+'-short-name']
#             }
#             try: grouped_traits[trait_key].append(missing)
#             except: grouped_traits[trait_key] = [missing]
    
for LANG in LANGS:
    print("### Generating EXEL files for", LANG)
    create_dir(TRAITS_DIR+"/"+LANG)
    create_dir(TRAITS_DIR+"/"+LANG+"/"+"NADs")
    create_dir(TRAITS_DIR+"/"+LANG+"/"+"4C")
    create_dir(TRAITS_DIR+"/"+LANG+"/"+"Others")
    create_dir(TRAITS_DIR+"/"+LANG+"/"+"Scores_19")
    create_dir(TRAITS_DIR+"/"+LANG+"/"+"Scores_HR")
    
    
    #for READER_TYPE in READER_TYPES:
    for key, trait_group in grouped_traits.items():
        BASE_PATH = TRAITS_DIR+'/'+LANG+'/'
        if key in NAD_NAMES: BASE_PATH+='NADs/'
        elif key in TRAITS_GROUP_4C: BASE_PATH+='4C/'
        elif key in TRAITS_GROUP_SCORES_19: BASE_PATH+='Scores_19/'
        elif key in TRAITS_GROUP_SCORES_HR: BASE_PATH+='Scores_HR/'
        else: BASE_PATH+='Others/'
        
        shortName = trait_group[0]['shortName'][LANG]
        if 'strong-and-weak-points-for-' in key: shortName = "Strong&Weak "+shortName
        file_path = BASE_PATH+shortName+" [id_"+key+"]"+'.xlsx'
        shutil.copyfile('./scripts/cognitive_v2/templates/trait.xlsx', file_path)
        workbook = load_workbook(file_path)
        
        
        # Rename first sheet
        sheet = workbook.active
        sheet.title = READER_TYPES[0]
        
        # Add next sheets for each reader type
        for READER_TYPE in READER_TYPES[1:]:
            sheet = workbook.active
            target = workbook.copy_worksheet(sheet)
            target.title = READER_TYPE
        
        # Add initial data
        for READER_TYPE in READER_TYPES:
            sheet = workbook[READER_TYPE]
            sheet["B2"] = key
            sheet["C2"] = READER_TYPE
            sheet["D2"] = READER_TYPES_DESCIPTIONS[READER_TYPE]
            sheet["C3"] = key# This will be replaced with name
        
        for trait in trait_group:
            sheet = workbook[trait['readerType']]
            sheet["B3"] = trait['abbreviation'][LANG]
            sheet["C3"] = trait['shortName'][LANG]
            sheet["B6"] = trait['shortDescription'][LANG]
            sheet["B7"] = trait['mainDefinition'][LANG]
            
            
            
            # Personalized texts - depends on the value
            for level in [1,2,3,4,5]:
                level_label = 'level_'+str(level)
                ###############################################################
                # Personalized descriptions ###################################
                ###############################################################
                personalized_descriptions = trait['descriptions'][level_label][LANG]
                for index, des in enumerate(personalized_descriptions):
                    row = 9+level
                    column = openpyxl.utils.cell.get_column_letter(2+3*index)
                    cell = column+str(row)
                    sheet[cell] = des
                ###############################################################
                # Personalized actions ########################################
                ###############################################################
                personalized_actions = trait['actions'][level_label][LANG]
                for index, act in enumerate(personalized_actions):
                    row = 16+level
                    column = openpyxl.utils.cell.get_column_letter(2+3*index)
                    cell = column+str(row)
                    sheet[cell] = act
                    
            ###############################################################
            # Highest and lowest values ###################################
            ###############################################################
            sheet['B24'] = trait['highestDefinition'][LANG]
            for index, act in enumerate(trait['actions']['highest'][LANG]):
                column = openpyxl.utils.cell.get_column_letter(2+3*index)
                sheet[column+'25'] = act
            # Lowest value
            sheet['B28'] = trait['lowestDefinition'][LANG]
            for index, act in enumerate(trait['actions']['lowest'][LANG]):
                column = openpyxl.utils.cell.get_column_letter(2+3*index)
                sheet[column+'29'] = act

                    
            ###############################################################
            # Neurobiological Effects ######################################
            ###############################################################
            for index, eff in enumerate(trait['neurobiologicalEffects'][LANG]):
                column = openpyxl.utils.cell.get_column_letter(2+3*index)
                sheet[column+'32'] = eff
            
        workbook.save(file_path)
            


