# Exporting data from existing initial database files
# into exel files matchin the to[_template


import copy
from openpyxl import load_workbook
import openpyxl.utils.cell
import shutil
from utils import *

import sys

PATH_TO_ELIA_ALGORITHMS  = '/home/adrian/Work/elia-algorithms' #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!<- SET THIS
sys.path.insert(1, PATH_TO_ELIA_ALGORITHMS+'/app/tips_detection_algorithm/v2/')
from algo import *



TIPS_DIR = BASE_DIR+'/tips'
recreate_dir(TIPS_DIR)


cognitiveTips = load_json_from_initial_database_file('cognitiveTips.js')


def get_traits(key):
    row = key.split('_')[0]
    TIP = next(t for t in ALL_TIPS if int(row) in t['rows'])
    return str(TIP['traits']).strip('[').strip(']')

for LANG in LANGS:
    print("### Generating EXEL files for", LANG)
    BASE_PATH = TIPS_DIR+'/'+LANG+'/'
    create_dir(BASE_PATH)
    file_path = BASE_PATH+'tips.xlsx'
    shutil.copyfile('./scripts/cognitive_v2/templates/tip.xlsx', file_path)
    workbook = load_workbook(file_path)
    
    
   # Rename first sheet
    sheet = workbook.active
    sheet.title = READER_TYPES[0]
    
    # Add next sheets for each reader type
    for READER_TYPE in READER_TYPES[1:]:
        sheet = workbook.active
        target = workbook.copy_worksheet(sheet)
        target.title = READER_TYPE
    for READER_TYPE in READER_TYPES:
        sheet = workbook[READER_TYPE]
        sheet["C2"] = READER_TYPE
        sheet["D2"] = READER_TYPES_DESCIPTIONS[READER_TYPE]
        
        matching_tips = [t for t in cognitiveTips if t['readerType']==READER_TYPE]
        matching_tips = sorted(matching_tips, key=lambda x: int(x['key'].split('_')[0]), reverse=False)
        for index, tip in enumerate(matching_tips):
            row = 5+(index*4)
            sheet.cell(row=row, column=2).value = tip['key']
            sheet.cell(row=row, column=3).value = tip['ageGroup']
            sheet.cell(row=row, column=4).value = tip['introduction'][LANG]
            sheet.cell(row=row, column=7).value = tip['text'][LANG]
            sheet.cell(row=row, column=10).value = tip['reasoning'][LANG]
            sheet.cell(row=row, column=13).value = get_traits(tip['key'])
        
        
        
            
    workbook.save(file_path)
            


